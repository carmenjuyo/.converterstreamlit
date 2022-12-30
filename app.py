import sys, os
from datetime import datetime
import calendar
import json
from streamlit_sortables import sort_items
from streamlit_tags import st_tags
from fuzzywuzzy import process
import pandas as pd
import streamlit as st
import openpyxl
import traceback
import gspread
import random
import string
from PIL import Image

#Import data connection modules
from modules.data_process import Gscon
from modules.data_ret import Gsret


st.set_page_config(page_title="Forecast Converter - JUYO", page_icon=Image.open('images/JUYOcon.ico'), layout="wide")

hide_default_format = """
       <style>
       footer {visibility: hidden;}
       header {visibility: hidden;}
       </style>
       """
st.markdown(hide_default_format, unsafe_allow_html=True)

def save_storage():
    
    with st.sidebar:

        with st.spinner('Storing data...'):
            
            gs_storage = []

            json_string = json.dumps(result_list)

            credentials = Gscon.run_credentials()

            gc = gspread.service_account_from_dict(credentials)

            sh = gc.open(st.secrets["private_gsheets_url"])

            for key, values in result_list.items():
                gs_storage.append(key)
                if(isinstance(values, list)):
                    for value in values:
                        gs_storage.append(value)
                else:
                    gs_storage.append(value)
            
            a=len(sh.sheet1.row_values(1)) + 65
            letter = chr(len(sh.sheet1.row_values(1)) + 65)

            start_row = 2
            end_row = start_row + len(gs_storage) - 1
            range1 = "%s%d:%s%d" % (letter, start_row, letter, end_row)
            
            cell_list = sh.sheet1.range(range1)

            sh.sheet1.update_cells(cell_list)

            result_str = ''.join(random.choice(string.ascii_letters) for i in range(8))

            z = 0
            for x in range(len(gs_storage)):
                cell_list[z].value = gs_storage[x]
                z = z + 1
            sh.sheet1.update_cells(cell_list)

            a = a - 64
            sh.sheet1.update_cell(1, a, f'{result_str}')

            st.write("""
                # Here you can check your input for future use.
                Save the generated password for later purposes.
                """)

            st.json(json_string, expanded=False)

        st.write(f'''
            #### Here is your password: 
            {result_str}
            ##### Save it well.
            ''')
        st.success('Input saved!')

def run_process(result_list):
    
    with st.spinner('runnning process...'):

        print(f"starting process...")

        # --- Declaring variables ---
        rn_c = 0
        rv_c = 0
        cRngn = []
        cRngv = []

        iSort_t = [] # Temp sorting list
        iSort_l = [] # For each loop sorting

        iDataRn = []
        iDataRnT = []
        iDataRv = []
        iDataRvT = []

        iMonths = []
        iDays = []
        sMonths = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Sept","Oct","Nov","Dec"]

        result_list['iSkipper'] = [round(float(i)) for i in result_list['iSkipper']]
        result_list['iStepper'] = [round(float(i)) for i in result_list['iStepper']]
        
        # Adding the selected sheets into a list, in that list, the months will be extracted of the sheets name.
        for x in cols: iMonths.append(x)
        for x in iMonths:
            str2Match = x

            highest = process.extractOne(str2Match,sMonths)
            highest = highest
            iDays.append(highest)

        # Set a list of the segments for later use in the sorting algo.

        for x in result_list['iSort']:
            iSort_t.append(x)
            iSort_l.append(x)
        
        # If no terminology is found, the script will end and show a warning
        if len(result_list['iTerm']) == 0:
            st.error('Err1: No terminology filled in. Press "enter" after typing the terminology!', icon='❌')
            return

        # Opens the Excel file from the client
        wb = openpyxl.load_workbook(uploaded_file_CLIENT, data_only=True)

        # Here the process of converting will start
        # With the use of Try, Except error messages will pop-up in case of mistakes.
        try:
            
            for x in range(len(iMonths)):
                
                ws = wb[(iMonths[x])]

                rn_c = 0
                rv_c = 0
                rn_c1 = 0
                rv_c1 = 0

                for z in sMonths:
                    try:
                        iDays[x].index(z)
                    except ValueError:
                        pass
                    else:
                        tMonth = z
                        if x == 0: eMonth = z

                print(f'Current month: {tMonth}')

                leap_year = [int(2024), int(2028), int(2032)]

                if tMonth == 'Jan' or tMonth == 'Mar' or tMonth == 'May' or tMonth == 'Jul' or tMonth == 'Aug' or tMonth == 'Oct' or tMonth == 'Dec': mDay = 31
                elif tMonth == 'Apr' or tMonth == 'Jun' or tMonth == 'Sep' or tMonth == 'Sept' or tMonth == 'Nov': mDay = 30
                elif tMonth == 'Feb' and st.session_state.year in leap_year: mDay = 29
                elif tMonth == 'Feb': mDay = 28
                else: mDay = 30

                # Here will begin the process of looking for the data in the currect sheet in the for loop
                # Because it can be stored in rows or columns, the for loop will run 2 times

                # //IDEA set variables for [0] or [1] (depending on row or column stored) to only have 2 for loops
                if result_list['iLoc'][0] == 'Rows':

                    for i in range(1, ws.max_row + 1):
                        for j in range(1, ws.max_column + 1):
                            if i == int(result_list['iLoc'][1]):
                                if result_list['iTerm'][0] == ws.cell(i,j).value:
                                    rn_c = rn_c + 1
                                    if int(rn_c) in set(result_list['iSkipper']):
                                        pass
                                    else:
                                        rn_c1 = rn_c1 + 1
                                        cRngn.append(f"{result_list['iTerm'][0]} {ws.cell(i,j)}")
                                        for row in ws.iter_rows(min_row=i + 1,max_row=i + mDay, min_col=j, max_col=j):
                                            for cell in row:
                                                iDataRn.append(cell.value)
                                        
                                        iDataRnT.append(iDataRn)
                                        iDataRn = []
                    
                    if rn_c1 != len(result_list['iSegments:']):

                        st.error(f"""
                            ##### Err2: ERROR for: {result_list['iTerm'][0]}. In total {len(result_list['iSegments:'])} segments entered. But {rn_c} segments were measured in the month / sheet: {iMonths[x]}.
                            See below an overview of the segments and their range that were succeeded:
                        """, 
                        icon="❌")
                        st.json(cRngn, expanded=True)
                        return

                    for i in range(1, ws.max_row + 1):
                        for j in range(1, ws.max_column + 1):
                            if i == int(result_list['iLoc'][1]):
                                if result_list['iTerm'][1] == ws.cell(i,j).value:
                                    rv_c = rv_c + 1
                                    if int(rv_c) in set(result_list['iStepper']):
                                        pass
                                        
                                    else:
                                        rv_c1 = rv_c1 + 1
                                        cRngv.append(f"{result_list['iTerm'][1]} {ws.cell(i,j)}")
                                        for row in ws.iter_rows(min_row=i + 1,max_row=i + mDay, min_col=j, max_col=j):
                                            for cell in row:
                                                iDataRv.append(round(cell.value,2))
                                        
                                        iDataRvT.append(iDataRv)
                                        iDataRv = []

                    if rv_c1 != len(result_list['iSegments:']):

                        st.error(f"""
                            ##### Err3: ERROR for: {result_list['iTerm'][1]}. In total {len(result_list['iSegments:'])} segments entered. But {rv_c} segments were measured in the month / sheet: {iMonths[x]}.
                            See below an overview of the segments and their range that were succeeded:
                        """, 
                        icon="❌")
                        st.json(cRngv, expanded=True)
                        return

                else:
                    
                    for i in range(1, ws.max_row + 1):
                        for j in range(1, ws.max_column + 1):
                            if j == int(result_list['iLoc'][1]):
                                if result_list['iTerm'][0] == ws.cell(i,j).value:
                                    rn_c = rn_c + 1
                                    if int(rn_c) in set(result_list['iSkipper']):
                                        pass
                                        
                                    else:
                                        rn_c1 = rn_c1 + 1
                                        cRngn.append(f"{result_list['iTerm'][0]} {ws.cell(i,j)}")
                                        for column in ws.iter_cols(min_row=i,max_row=i, min_col=j + 1, max_col=j + mDay):
                                            for cell in column:
                                                iDataRn.append(cell.value)
                                        
                                        iDataRnT.append(iDataRn)
                                        iDataRn = []

                    if rn_c1 != len(result_list['iSegments:']):
                        
                        st.error(f"""
                            ##### ERROR for: {result_list['iTerm'][0]}. In total {len(result_list['iSegments:'])} segments entered. But {rn_c} segments were measured in the month / sheet: {iMonths[x]}.
                            See below an overview of the segments and their range that were succeeded:
                        """, 
                        icon="❌")
                        st.json(cRngn, expanded=True)
                        return

                    for i in range(1, ws.max_row + 1):
                        for j in range(1, ws.max_column + 1):
                            if j == int(result_list['iLoc'][1]):
                                if result_list['iTerm'][1] == ws.cell(i,j).value:
                                    rv_c = rv_c + 1
                                    if int(rv_c) in set(result_list['iStepper']):
                                        pass

                                    else:
                                        rv_c1 = rv_c1 + 1
                                        cRngv.append(f"{result_list['iTerm'][1]} {ws.cell(i,j)}")
                                        for column in ws.iter_cols(min_row=i,max_row=i, min_col=j + 1, max_col=j + mDay):
                                            for cell in column:
                                                iDataRv.append((cell.value))
                                        
                                        iDataRvT.append(iDataRv)
                                        iDataRv = []

                    if rv_c1 != len(result_list['iSegments:']):
                        
                        st.error(f"""
                            ##### ERROR for: {result_list['iTerm'][1]}. In total {len(result_list['iSegments:'])} segments entered. But {rv_c} segments were measured in the month / sheet: {iMonths[x]}.
                            See below an overview of the segments and their range that were succeeded:
                        """, 
                        icon="❌")
                        st.json(cRngv, expanded=True)
                        return

            # When all the data is stored in a 2d array, it needs to be sorted, as the data is stored as it was presented in the sheet
            # So now it will loop through all the months, segments, and non-sorted segments.
            # When a match is found the 2d array will be resorted in the correct order
            a = 0
            for x in range(len(iMonths)):
                
                for i in range(len(result_list['iSegments:'])):

                    strFind = result_list['iSegments:'][i]

                    for y in range(len(iSort_l)):

                        strStored = iSort_l[y]

                        #print(f"L: {i} - {(result_list['iSegments:'][i])} | R: {y} - {(iSort_l[y])}")

                        if strFind == strStored:
                            
                            if i == y:
                                #print(f"-- Level & Name match: {strFind} = {i}")
                                pass
                            else:
                                #print(f"- Name match: {strFind} : {i} - {y}")
                                #print(f"- reordering...")
                                
                                arrTemp = iDataRnT[y + a]
                                arrTempV = iDataRvT[y + a]

                                iDataRnT[y + a] = iDataRnT[i + a]
                                iDataRvT[y + a] = iDataRvT[i + a]

                                iDataRnT[i + a] = arrTemp
                                iDataRvT[i + a] = arrTempV

                                temp = iSort_l[i]

                                iSort_l[i] = result_list['iSegments:'][i]
                                iSort_l[y] = temp
                
                if x == 0:
                    a = len(iSort_l)
                else:
                    a = a + len(iSort_l)

                iSort_l.clear()
                for z in iSort_t: iSort_l.append(z)

            # Here will be the two 2d array merged together for later purposes
            # The array will then be put into an dataframe so it can be transposed
            d =[]

            for x in range(len(iDataRnT)):
                d.append(iDataRnT[x])
                d.append(iDataRvT[x])

            df2 = pd.DataFrame(data=d)

            df2 = df2.T
            
            # Create an new Excel file for the output data
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title='sheet0'

            # Here will the columns of the Juyo file be stored
            y = 1
            for x in range(len(all_columns)):
                sheet.cell(row=1, column=y).value=all_columns[x]
                y = y + 1

            # the next process is for storing the data in the correct order for the correct months
            # Because the array is already sorted correctly due to the sorting process, it now only has to be put nice together
            # 

            x = 2 # COLUMN
            y = 2 # ROW
            t = 2 # ROW
            s = 1 # SEGMENTS

            for i in range(len(iDataRnT)):

                for z in range(len(iDataRnT[i])):

                    sheet.cell(row=y, column=x).value=iDataRnT[i][z]
                    if result_list['iDataSt'] == "ADR":
                        sheet.cell(row=y, column=x+1).value=(iDataRnT[i][z] * iDataRvT[i][z])
                    else:
                        sheet.cell(row=y, column=x+1).value=iDataRvT[i][z]
                    y = y + 1

                if s == len(result_list['iSegments:']):
                    x = 2
                    s = 1
                    if i <= len(result_list['iSegments:']):
                        t = 2 + (len(iDataRnT[i]))
                    else:
                        t = t + (len(iDataRnT[i]))
                else:
                    s = s + 1
                    x = x + 2
                    y = t 

            # Here is why the user needed to select the starting year. Here it will look at the data and store the dates
            datetime_object = datetime.strptime(eMonth, "%b")
            month_number = datetime_object.month

            datelist = pd.date_range(datetime(st.session_state.year, month_number, 1), periods=sheet.max_row - 1).to_pydatetime().tolist()
            
            i = 2
            for x in range(len(datelist)):
                sheet.cell(row=i, column=1).value=datelist[x]
                i = i + 1

            for cell in sheet["A"]:
                cell.number_format = "YYYY/MM/DD"
            
            # Temp save the data to an excel file, otherwise it wouldn't work with downloading the excel file later.
            wb.save('NameFile.xlsx')

            df3 = pd.read_excel('NameFile.xlsx')

            st.dataframe(df3)      
        
        except Exception as e:
            
            traceback.print_exc()

            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]

            credentials = Gscon.run_credentials()

            gc = gspread.service_account_from_dict(credentials)

            sh = gc.open(st.secrets["private_gsheets_url_log"])
            a=len(sh.sheet1.col_values(1))

            sh.sheet1.update_cell(a + 1, 1, f'Err4: {exc_type}; {exc_obj}; ({str(e)}), line: {exc_tb.tb_lineno}, in {fname} {datetime.utcnow()}')
            st.error(f'Err4: {exc_type}; {exc_obj}; ({str(e)}), line: {exc_tb.tb_lineno}, in {fname} | UTC: {datetime.utcnow()}')
            return

    st.success('Process ran!')

    with open("NameFile.xlsx", "rb") as file:
        st.download_button(
            label="Click me to download excel!",
            data=file,
            file_name=f'{uploaded_file_JUYO.name}',
            mime="application/octet-stream"
            )

# Header of the page
with st.container():

    l_column, m_column, r_column = st.columns([1,6,1])

    with m_column:
        st.markdown("<h1 style='text-align: center; color: white;'>♾️Forecast / budget converter</h1>", unsafe_allow_html=True)
    with r_column:
        st.image(Image.open('images/JUYO3413_Logo_Gris1.png'))

# Input part of the page
with st.container():

    st.write("---")

    disabled = 1

    # --- Initialising SessionSate ---
    if "dict" not in st.session_state:
        st.session_state.dict = {}

    if "year" not in st.session_state:
        st.session_state.year = 0

    left_column, right_column = st.columns(2)

    with left_column:
        
        st.markdown("<h2 style='text-align: center; color: white;'>Forecast / budget file client</h2>", unsafe_allow_html=True)

        uploaded_file_CLIENT = st.file_uploader("Upload file client", type=".xlsx")

        if uploaded_file_CLIENT:
            
            st.markdown("### Select wanted sheets in chronological order for conversion.")

            tabs = pd.ExcelFile(uploaded_file_CLIENT).sheet_names

            cols = st.multiselect('Select sheets in the **chronological** order of the months:', tabs)
            
            if len(cols) == 0: st.warning('Please press enter after each input.')
            st.write('You selected:', len(cols))

    with right_column:
        
        st.markdown("<h2 style='text-align: center; color: white;'>Format file JUYO</h2>", unsafe_allow_html=True)
        
        uploaded_file_JUYO = st.file_uploader("Upload JUYO file", type=".xlsx")

        if uploaded_file_JUYO:
                
            df = pd.read_excel(uploaded_file_JUYO)

            all_columns = df.columns
            shape1 = df.shape

            df = df[[k for i, k in enumerate(df.columns, 0) if i % 2 != 0]]

            shape = df.shape
            st.markdown(f"<p style='text-align: center; color: white;'>{shape[1]} segments detected in Juyo file</p>", unsafe_allow_html=True)
    
    if uploaded_file_JUYO and uploaded_file_CLIENT and len(cols) > 0:
        
        st.write("---")
        
        st.markdown("<h1 style='text-align: left; color: white;'>Enter password to retreive previous data stored:</h1>", unsafe_allow_html=True)
        st.markdown("<h6 style='text-align: left; color: white;'>When the password is entered, it will retreive the data stored, so that you won't have to enter all the input</h6>", unsafe_allow_html=True)

        use_password = st.radio(
                label="-",
                options=("Use password", "Enter new input"),
                horizontal=True,
                index=0,
                label_visibility='collapsed'
            )
        
        if use_password == 'Use password':

            form = st.form('password_form', clear_on_submit=False)
            key_s = form.text_input('Enter password:')

            if form.form_submit_button("Submit"):
                result_list = Gsret.retreive_data(key_s)
                st.session_state.dict = result_list
            
            # //BUG displays 2 times success, raises because of imported module (29/12/2022 17:16 can be fixed after merge)
            if not bool(st.session_state.dict) == False:
                st.success(f'password: {key_s} succesfull', icon='✅')
                st.json(st.session_state.dict, expanded=False)    
            else: pass

            if bool(st.session_state.dict) == False:
                disabled = 1
            else:
                disabled = 0

            st.write('## Select starting year of first sheet.')

            st.session_state.year = st.select_slider(
                label="# .",
                options=range(datetime.today().year - 2, datetime.today().year + 3),value=datetime.today().year,
                label_visibility='collapsed',
                disabled=disabled
                )

            if st.button("Start converting process.", key="run1", disabled=disabled): 
                run_process(st.session_state.dict)
        
        elif use_password == 'Enter new input':

            st.write("---")

            keywords = st_tags(
                    label="""
                        # Enter your segments in the exact order they appear in your Excel file:
                        So go to your sheet to where the segments with data is stored, 
                        and then go from left to right or top to bottom and then fill in all your segments 
                        here in that order.
                            """,
                    text='Press enter to add more',
                    suggestions=['leisure', 'Leisure', 'groups', 
                                'Groups', 'group', 'Group', 
                                'Business', 'business', 'corporate',
                                'Direct', 'direct', 'Indirect', 'indirect'
                                'individual', 'packages', 'complementary', 'house'
                                ],
                    maxtags = shape[1],
                    key='1')

            expander =  st.expander('See explantion of why the segments needs to be in chronological order segments :')
            
            with open(f'content/explanation1.md', 'r') as f:
                    expander.markdown(f.read())

            st.markdown(f"<h3 style='text-align: center; color: white;'>{len(keywords)} segments of {shape[1]} entered</h3>", unsafe_allow_html=True)

            if len(keywords) == shape[1] - 1:
                st.warning('You are 1 segment short! You can continue with 1 segment short, but the last segments in Juyo will be kept empty', icon='⚠️')
            elif len(keywords) < shape[1]:
                st.warning(f'You are {shape[1] - len(keywords)} segments short! Please add more segments to match the Juyo segments',icon='❌')

            left_column, right_column = st.columns(2)

            with left_column:
            
                st.markdown(f"<h3 style='text-align: center; color: white;'>Juyo active segments in correct order</h3>", unsafe_allow_html=True)
                sorted_items1 = sort_items(df.columns.to_list(), direction='vertical')

            with right_column:
                
                st.markdown(f"<h3 style='text-align: center; color: white;'>Map segments to match exact with Juyo</h3>", unsafe_allow_html=True)
                sorted_items2 = sort_items(keywords, direction='vertical')

            st.write('---')

            st.write("# Is the data stored as *room nights and revenue* or *room nights and ADR*")
            option = st.radio(
                    label="## .",
                    options=('Room nights & Revenue', 'Room nights & ADR'),
                    label_visibility='collapsed')

            if option == 'Room nights & Revenue':
                term = "revenue"
                iData_choice = ["Rev"]
            else:
                term = 'ADR'
                iData_choice = ["ADR"]
            
            st.write(f"""
                ## Enter the terminology used in Excel file for room nights and {term}:
                For example; roonnights = Rms, Rn, etc. Revenue = Rev, Rvu, etc. (press ENTER when filled in!)
                """)
            
            terminology = []

            terminologyR = st_tags(
                label='Enter the terminology of **room nights:**',
                text='Press enter when terminology of room nights is entered !',
                suggestions=['rn', 'RN', 'Rn', 
                            'Rev', 'REV', 'rev', 
                            'ADR', 'Adr', 'adr',
                            ],
                maxtags = 1,
                key='t1')

            terminologyR1 = st_tags(
                label=f'Enter the terminology of **{term}:**',
                text=f'Press enter when terminology of {term} is entered !',
                suggestions=['rn', 'RN', 'Rn', 
                            'Rev', 'REV', 'rev', 
                            'ADR', 'Adr', 'adr',
                            ],
                maxtags = 1,
                key='t2')
            
            terminology = terminologyR + terminologyR1

            st.write(f''' ### Are the headers of the terminology room nights and {term} stored in a row or in a column?''')
            storage = st.radio(
                label='x',
                options=('Rows', 'Columns'),
                label_visibility='collapsed')

            with st.expander('Click for more explantion'):
                
                st.write(f'Here you can see an example why we ask for the row or column. E.g.:')
                image = Image.open('images/voorbeeld_excel.png')
                st.image(image)
                with open(f'content/explanation2.md', 'r') as f:
                    st.markdown(f.read())

            if storage == 'Rows':
                
                st.write(f''' ### in which row can the terminology of room nigts and {term} be found?"''')

                row_n = st.text_input(f"x", label_visibility='collapsed')
            
            elif storage == 'Columns':

                st.write(f''' ### in which row can the terminology of room nigts and {term} be found?"''')

                row_n = st.text_input(f"x", label_visibility='collapsed').lower()
                try: row_n = ord(row_n) - 96
                except: pass
            
            st.write('### Skip terminology on certain places?')
            skip_term = st.checkbox('Want to skip terminology on certain places?')

            with st.expander('Click for more explantion'):
            
                st.write(f'You have just indicated the terminology (room nights and {term}) is in {storage} {row_n}. It may happen that on that same {storage} the terminology (room nights and {term}) of totals are included, but should not be included. E.g.:')

                st.image(Image.open('images/voorbeeld_excel.png'))
                
                with open(f'content/explanation3.md', 'r') as f:
                    st.markdown(f.read())

            if skip_term:

                l_c, r_c = st.columns(2) 

                with l_c:
                    Skipper = st_tags(
                        label=f'### Skip a terminology in order {terminologyR}',
                        text='Press enter to add more',
                        suggestions=['1', '2', '3', 
                                    '4', '5', '6'],
                        maxtags = 10,
                        key='3')

                with r_c:
                    Skipper1 = st_tags(
                        label=f'### Skip a terminology in order {terminologyR1}',
                        text='Press enter to add more',
                        suggestions=['1', '2', '3', 
                                    '4', '5', '6'],
                        maxtags = 10,
                        key='4')  

            else:
                Skipper = []
                Skipper1 = []

            st.write('---')

            st.write('# Select starting year of first sheet')
            st.write('This is needed, as the script needs to know at which year the first year is starting')
            st.session_state.year = st.select_slider(
                label='x.',
                label_visibility='collapsed',
                options=range(datetime.today().year - 2, datetime.today().year + 3),value=datetime.today().year)

            st.write('---')

            st.markdown(f"<h1 style='text-align: center; color: white;'>-- Please check if everything is correct! --</h1>", unsafe_allow_html=True)
            
            with st.expander('Click for more explantion'):
                with open(f'content/explanation4.md', 'r') as f:
                    st.markdown(f.read())

            result_list = {
                'iSegments:': sorted_items2,
                'iTerm': terminology,
                'iSort': keywords,
                'iSkipper': [round(float(i)) for i in Skipper],
                'iStepper': [round(float(i)) for i in Skipper1],
                'iDataSt': iData_choice,
                'iLoc': [storage, row_n]
            }

            st.write('Click to expand the data: 👇')
            st.json(result_list, expanded=False)

            st.write('---')
            st.markdown(' ## Checked the data and the data is correct?')
            if st.checkbox('Data is correct'):
                if len(keywords) < shape[1] - 1:
                    st.warning(f'''{len(keywords)} segments entered, but {shape[1]} segments active in Juyo.
                    Please match the number of segments with Juyo or be 1 segment short.''')
                    disabled = 1

                else:
                    disabled = 0
                    st.info('If it is the first time, it is wise before storing the data, to first run the process to check if the data output is correct.', icon='ℹ️')

            if st.button('Store the data for next time', key="store", disabled=disabled): 
                save_storage()

            if st.button("Start converting process", key="run4", disabled=disabled):                        
                run_process(result_list)
